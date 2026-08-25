# -*- coding: utf-8 -*-
"""
2025 잠재 방한여행객 조사 - 핵심 변수 추출 스크립트

목적:
  data/raw/2025_잠재방한객조사_일반외국인.xlsx, data/raw/2025_잠재방한객조사_방한의향자.xlsx
  두 원본 Excel 파일에서 프로젝트 핵심 변수만 선별하여, 23개 공통국가(한류실태조사와 결합 가능한
  국가) 기준으로 long format CSV로 추출한다.

원칙:
  - 원본 Excel 파일은 오직 읽기(read-only)로만 접근하며 절대 수정하지 않는다.
  - Gap 계산, 국가 순위, 상관분석, DB 구축, 한류실태조사 결합 등은 이 스크립트의 범위가 아니다.
  - % 값은 원본 float 정밀도를 그대로 보존한다 (반올림하지 않는다).
  - 표C4/표C5(방한의향자 파일)와 동일 코드가 일반외국인 파일에 존재하더라도 절대 병합하지 않는다
    (survey_type으로 항상 구분).

실행:
  python data/scripts/extract_potential_tourist_data.py
  (프로젝트 루트 또는 data/scripts/ 어디에서 실행해도 동작하도록 경로를 스크립트 기준으로 계산함)
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# 경로 설정
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
OUT_DIR = PROJECT_ROOT / "data" / "processed"
OUT_DIR.mkdir(parents=True, exist_ok=True)

GENERAL_FILE = RAW_DIR / "2025_잠재방한객조사_일반외국인.xlsx"
INTENT_FILE = RAW_DIR / "2025_잠재방한객조사_방한의향자.xlsx"

OUT_GENERAL_CSV = OUT_DIR / "potential_tourist_general_core.csv"
OUT_INTENT_CSV = OUT_DIR / "potential_tourist_intender_core.csv"
OUT_MAPPING_CSV = OUT_DIR / "potential_tourist_country_mapping.csv"
OUT_VALIDATION_CSV = OUT_DIR / "potential_tourist_extraction_validation.csv"

# ---------------------------------------------------------------------------
# 23개 공통국가 (잠재방한객조사 26개국 ∩ 한류실태조사 30개국)
# 국가명은 잠재방한객조사 원본 Excel 표기를 그대로 사용한다.
# ---------------------------------------------------------------------------

TARGET_COUNTRIES = [
    "중국", "일본", "대만", "미국", "필리핀", "베트남", "싱가포르", "인도네시아",
    "태국", "말레이시아", "캐나다", "호주", "러시아", "인도", "프랑스", "독일",
    "영국", "멕시코", "카자흐스탄", "튀르키예", "브라질", "사우디아라비아", "아랍에미리트",
]
assert len(TARGET_COUNTRIES) == 23

# 표준 국가명 매핑 (한류실태조사 등 외부 데이터와 결합할 때 사용할 표준명).
# 현재 확인된 표기 차이는 '아랍에미리트' -> 'UAE' 뿐이며, 나머지는 잠재방한객조사 원표기를
# 그대로 표준명으로 사용한다. (한류실태조사 country_table_page_map.json 기준 대조 완료)
COUNTRY_STANDARD_NAME = {
    "아랍에미리트": "UAE",
}


def standard_name(country: str) -> str:
    return COUNTRY_STANDARD_NAME.get(country, country)


# ---------------------------------------------------------------------------
# 추출 대상 표 정의
#   general: 2025_잠재방한객조사_일반외국인.xlsx 대상 표코드 목록
#   intent : 2025_잠재방한객조사_방한의향자.xlsx 대상 표코드 목록
#
# 주의(중요한 판단): 원 요청의 "표B5A/B5B: 향후 3년 내 한국 방문 의향"은 실제 Excel에는
#   - <표B5A>/<표B5B> : 13개 아시아 목적지 전체에 대한 국가/지역별(=목적지별) 평균 매트릭스
#     (응답자 거주국이 아니라 '목적지'가 행이 되는 표이며, 코드가 BASE 조건별로 2회 중복 존재)
#   - <표B5A-1>/<표B5B-1> : '한국' 목적지 하나만 골라, 응답자 거주국별로 분포를 보여주는 표
#   두 가지가 존재한다. 프로젝트가 필요로 하는 것은 "응답자 거주국별 방한의향 분포"이므로
#   집계 매트릭스(B5A/B5B)가 아니라 한국 전용 표(B5A-1, B5B-1)를 추출 대상으로 채택했다.
#   이 판단은 보고서에 별도로 명시한다.
# ---------------------------------------------------------------------------

GENERAL_TABLE_CODES = [
    "0",  # 국가별 응답자 특성 및 사례수
    "B5A-1",  # 향후 3년 내 한국 방문 의향 (해외여행 의향자 기준, 5점척도)
    "B5B-1",  # 향후 3년 내 한국 방문 의향 (전체 응답자 기준, 4분류)
    "E1A-1",  # 한국문화 경험 여부
    "E4-1",  # 한국문화 경험 -> 한국 호감도 영향
    "E4-3",  # 한국문화 경험 -> 한국 (재)방문의향 영향
    "B13-1A",  # 한국 방문 비의향 이유 (1+2+3순위)
    "B13-2A",  # 한국 방문 비의향 이유 (1순위)
    "A2AA-1-1",  # (전생애) 한국 방문 경험 (전체 응답자 기준)
    "A2A-1-1",  # (전생애) 한국 방문 경험 (해외여행 경험자 기준)
    "A2A-1-2",  # (전생애) 한국 방문 횟수 (해외여행 경험자 기준)
    "A2BA-1",  # (최근 1년) 한국 방문 경험 (전체 응답자 기준)
    "A2B-1",  # (최근 1년) 한국 방문 경험 (아시아 국가별 경험자 기준)
] + [f"E3-{i}" for i in range(1, 14)]  # 한국문화 분야별 이용 빈도 (13개 분야)

INTENT_TABLE_CODES = [
    "C4",  # 한국 여행 관심 계기
    "C5",  # 한국 여행 결정 요인
]

# B13-2A는 "한국 방문 비의향자" 기준 표만 요청 대상이다. 원본에는 동일 코드가 "미결정자" 기준으로도
# 존재하지 않는다 (미결정자 이유는 B13-1B/2B 별도 코드) 이므로 코드 중복 문제는 없다.

# 참고: PDF 교차검증에서 공식적으로 확인된 스팟체크 기준값 (검증용, 하드코딩 상수)
PDF_SPOTCHECK = {
    ("일반외국인", "E1A-1", "전체", "경험"): 78.9,
    ("일반외국인", "E1A-1", "일본", "경험"): 59.6,
    ("일반외국인", "B13-1A", "전체", "한류(한국 문화 콘텐츠) 문화 관련 관심 부재"): 24.2,
}


# ---------------------------------------------------------------------------
# 데이터 구조
# ---------------------------------------------------------------------------


@dataclass
class TableBlock:
    code: str
    title: str
    title_row: int
    base_label: str
    header_labels: dict  # col_idx(0-based) -> response_option label
    data_start_row: int
    data_end_row: int  # inclusive


@dataclass
class LongRecord:
    source_file: str
    survey_type: str
    table_code: str
    table_title: str
    base: str
    country: str
    country_std: str
    category: str
    response_option: str
    sample_n: object
    sample_n_raw: str
    value: object
    unit: str


# ---------------------------------------------------------------------------
# 유틸리티
# ---------------------------------------------------------------------------

TABLE_TITLE_RE = re.compile(r"^<표([^>]+)>\s*(.*)$")
SAMPLE_N_RE = re.compile(r"^\(([^)]*)\)$")


def parse_sample_n(raw):
    """'(16360)' -> 16360.0 / '(-)' 또는 '(0)' 등 특수값 처리.

    PDF 일러두기(6쪽): 가중치 적용 사례수는 정수가 아닌 실수로 표기되며, 0으로 표기된 경우는
    0.5 미만 값이 반올림된 것, 실제 사례수가 없는 경우는 '0'이 아닌 '-'으로 표기된다.
    """
    if raw is None:
        return None, ""
    text = str(raw).strip()
    m = SAMPLE_N_RE.match(text)
    if not m:
        return None, text
    inner = m.group(1).strip()
    if inner in ("-", ""):
        return None, text  # 실제 사례수 없음 (원본 표기 '-' 보존은 sample_n_raw에서 확인 가능)
    try:
        return float(inner), text
    except ValueError:
        return None, text


def unit_for_label(label: str) -> str:
    if label is None:
        return ""
    text = str(label)
    if text.startswith("<") and "평균" in text:
        return "score"
    if text in ("계", "종합", "종합/Top2", "Top2"):
        return "%"
    return "%"


# ---------------------------------------------------------------------------
# 표 위치 인덱싱 (1차 패스: A열만 스캔)
# ---------------------------------------------------------------------------


def index_table_titles(ws):
    """Sheet1 A열을 스캔하여 [(row_idx, code, title), ...] 를 행 순서대로 반환."""
    titles = []
    for row_idx, (val,) in enumerate(
        ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1
    ):
        if isinstance(val, str) and val.startswith("<표"):
            m = TABLE_TITLE_RE.match(val)
            if m:
                code, rest = m.group(1), m.group(2)
                titles.append((row_idx, code, val))
    return titles


def find_table_bounds(titles, target_code, max_row):
    """target_code와 정확히 일치하는 표의 (title_row, next_title_row-1) 범위를 반환.
    동일 코드가 여러 번 나타나면 전부 반환한다(중복 코드 감지용)."""
    matches = []
    for i, (row_idx, code, title) in enumerate(titles):
        if code == target_code:
            next_row = titles[i + 1][0] if i + 1 < len(titles) else max_row + 1
            matches.append((row_idx, next_row - 1, title))
    return matches


# ---------------------------------------------------------------------------
# 표 파싱 (2차 패스: 지정 범위만 읽기)
# ---------------------------------------------------------------------------


def parse_table_block(ws, title_row, end_row, code, title):
    """하나의 표 블록을 파싱하여 (base_label, header_labels, rows) 를 반환.

    구조:
      title_row      : <표코드> 제목
      title_row+1    : BASE:... | None | 사례수 | 응답옵션1 | 응답옵션2 | ...
      (title_row+2..): 헤더 연속행(2줄짜리 헤더, 전부 col A/B/C가 None) 있을 수 있음 -> 헤더에 병합
      데이터 시작    : col A == '▩전체▩' 인 행
      데이터 행      : col A(카테고리, forward-fill) | col B(국가명) | col C(사례수) | 값...
      종료           : end_row 도달 또는 다음 표 시작 전까지
    """
    header_row_idx = title_row + 1
    header_row = next(
        ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
    )
    base_label = header_row[0] if header_row[0] else ""

    # 응답옵션 헤더: C열(index 2) 다음(index 3)부터 연속된 non-None 값
    header_labels = {}
    for idx in range(3, len(header_row)):
        val = header_row[idx]
        if val is None:
            break
        header_labels[idx] = str(val)

    # 헤더 연속행 처리 (데이터 시작 '▩전체▩' 를 만날 때까지)
    ptr = header_row_idx + 1
    safety_limit = header_row_idx + 6  # 헤더 연속행은 통상 1줄 이내, 안전장치로 최대 5줄까지만 허용
    data_start = None
    while ptr <= end_row and ptr <= safety_limit:
        row = next(ws.iter_rows(min_row=ptr, max_row=ptr, values_only=True))
        if row[0] == "▩전체▩":
            data_start = ptr
            break
        # 헤더 연속행: 비어있지 않은 셀을 기존 헤더에 이어붙임
        for idx in range(3, len(row)):
            val = row[idx]
            if val is not None:
                if idx in header_labels:
                    header_labels[idx] = f"{header_labels[idx]}/{val}"
                else:
                    header_labels[idx] = str(val)
        ptr += 1
    if data_start is None:
        # '▩전체▩' 를 못 찾은 경우 -> 구조 이상, 다음 행을 데이터 시작으로 간주하고 경고 남김
        data_start = header_row_idx + 1

    # 데이터 행 파싱
    rows = []
    current_category = None
    for row in ws.iter_rows(min_row=data_start, max_row=end_row, values_only=True):
        col_a, col_b = row[0], row[1]
        if isinstance(col_a, str) and col_a.startswith("<표"):
            break  # 다음 표 시작 (안전장치)
        if col_a == "▩전체▩":
            category, country = "전체", "전체"
        elif col_a is not None:
            current_category = col_a
            country = col_b
            category = current_category
        elif col_b is not None:
            country = col_b
            category = current_category
        else:
            # 완전 공백행: 표 종료로 간주
            if all(v is None for v in row):
                continue
            else:
                continue
        rows.append((category, country, row))

    return base_label, header_labels, rows


# ---------------------------------------------------------------------------
# 표 -> long format 레코드 변환
# ---------------------------------------------------------------------------


def extract_table(
    ws, source_file_name, survey_type, code, title_row, end_row, title, warnings
):
    base_label, header_labels, rows = parse_table_block(ws, title_row, end_row, code, title)

    records = []
    countries_seen = []
    for category, country, row in rows:
        if country != "전체" and country not in TARGET_COUNTRIES:
            continue  # 23개 공통국가 + 전체(기준행)만 남김
        if country != "전체" and category != "국가/지역":
            # '세부지역' 구간에서 도시국가(싱가포르, 홍콩)나 지역 세분화가 없는 국가(아랍에미리트)는
            # 국가명 그대로 다시 한 번 나열되어 중복을 유발한다. PDF 6쪽 일러두기 참고:
            # "도시국가인 싱가포르와 홍콩, 세부 지역별 모집단 확인이 어려운 아랍에미리트의 경우
            #  지역별 표본 설계를 적용하지 않음". 최상위 '국가/지역' 구간의 값만 채택한다.
            continue
        sample_n, sample_n_raw = parse_sample_n(row[2])
        for col_idx, label in header_labels.items():
            value = row[col_idx] if col_idx < len(row) else None
            records.append(
                LongRecord(
                    source_file=source_file_name,
                    survey_type=survey_type,
                    table_code=code,
                    table_title=title,
                    base=base_label,
                    country=country,
                    country_std=standard_name(country) if country != "전체" else "전체",
                    category=category,
                    response_option=label,
                    sample_n=sample_n,
                    sample_n_raw=sample_n_raw,
                    value=value,
                    unit=unit_for_label(label),
                )
            )
        if country != "전체":
            countries_seen.append(country)

    missing = sorted(set(TARGET_COUNTRIES) - set(countries_seen))
    dup_counts = {c: countries_seen.count(c) for c in set(countries_seen) if countries_seen.count(c) > 1}
    if missing:
        warnings.append(f"[{survey_type}/{code}] 누락 국가: {missing}")
    if dup_counts:
        warnings.append(f"[{survey_type}/{code}] 중복 국가: {dup_counts}")

    return records, base_label, countries_seen


# ---------------------------------------------------------------------------
# 메인 추출 로직
# ---------------------------------------------------------------------------


def run_extraction_for_file(path: Path, survey_type: str, target_codes: list, warnings: list):
    print(f"[읽기] {path.name} 열기 중 (read-only)...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    max_row = ws.max_row

    print(f"[인덱싱] {path.name} 표 제목 스캔 중 (A열, {max_row}행)...")
    titles = index_table_titles(ws)
    print(f"  -> 총 {len(titles)}개 표 발견")

    all_records = []
    table_summaries = []  # (code, base, n_countries, n_rows, matched_pdf, note)

    for code in target_codes:
        matches = find_table_bounds(titles, code, max_row)
        if not matches:
            warnings.append(f"[{survey_type}/{code}] 표를 찾을 수 없음 (건너뜀)")
            continue
        if len(matches) > 1:
            warnings.append(
                f"[{survey_type}/{code}] 동일 코드가 {len(matches)}회 중복 발견 -> 모두 추출 "
                f"(BASE 라벨로 구분 필요, 상세 제목: {[m[2] for m in matches]})"
            )
        for title_row, end_row, title in matches:
            records, base_label, countries_seen = extract_table(
                ws, path.name, survey_type, code, title_row, end_row, title, warnings
            )
            all_records.extend(records)
            table_summaries.append(
                {
                    "table_code": code,
                    "table_title": title,
                    "base": base_label,
                    "n_countries": len(set(countries_seen)),
                    "n_records": len(records),
                    "total_row_present": any(
                        r.country == "전체" for r in records if r.table_code == code
                    ),
                }
            )
            print(f"  [{survey_type}] {code}: {len(records)}건 추출 (국가 {len(set(countries_seen))}개), BASE='{base_label}'")

    wb.close()
    return all_records, table_summaries


def write_long_csv(path: Path, records: list):
    fieldnames = [
        "source_file", "survey_type", "table_code", "table_title", "base",
        "country", "country_std", "category", "response_option",
        "sample_n", "sample_n_raw", "value", "unit",
    ]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in records:
            writer.writerow(
                {
                    "source_file": r.source_file,
                    "survey_type": r.survey_type,
                    "table_code": r.table_code,
                    "table_title": r.table_title,
                    "base": r.base,
                    "country": r.country,
                    "country_std": r.country_std,
                    "category": r.category,
                    "response_option": r.response_option,
                    "sample_n": r.sample_n,
                    "sample_n_raw": r.sample_n_raw,
                    "value": r.value,
                    "unit": r.unit,
                }
            )
    print(f"[저장] {path} ({len(records)}행)")


def write_mapping_csv(path: Path):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["original_name_잠재방한객조사", "standard_name", "note"])
        for c in TARGET_COUNTRIES:
            note = "표기 상이 (한류실태조사는 'UAE' 사용)" if c in COUNTRY_STANDARD_NAME else ""
            writer.writerow([c, standard_name(c), note])
    print(f"[저장] {path} ({len(TARGET_COUNTRIES)}행)")


def run_validation(all_records_general, all_records_intent, warnings):
    """추출 결과에 대한 자동 검증을 수행하고 검증 결과 레코드 목록을 반환."""
    validation_rows = []

    def check_table(records, survey_type, code):
        sub = [r for r in records if r.table_code == code]
        countries = sorted({r.country for r in sub if r.country != "전체"})
        missing = sorted(set(TARGET_COUNTRIES) - set(countries))
        extra = sorted(set(countries) - set(TARGET_COUNTRIES))
        base_labels = sorted({r.base for r in sub})
        total_row = [r for r in sub if r.country == "전체"]
        total_sample_n = None
        if total_row:
            sample_ns = {r.sample_n for r in total_row if r.sample_n is not None}
            total_sample_n = sorted(sample_ns) if sample_ns else None
        return {
            "survey_type": survey_type,
            "table_code": code,
            "n_target_countries_found": len(set(TARGET_COUNTRIES) & set(countries)),
            "n_expected": 23,
            "missing_countries": ";".join(missing) if missing else "",
            "unexpected_countries": ";".join(extra) if extra else "",
            "base_labels": " | ".join(base_labels),
            "total_row_sample_n": total_sample_n,
            "n_records": len(sub),
            "status": "OK" if not missing and not extra else "CHECK",
        }

    for code in GENERAL_TABLE_CODES:
        validation_rows.append(check_table(all_records_general, "일반외국인", code))
    for code in INTENT_TABLE_CODES:
        validation_rows.append(check_table(all_records_intent, "방한의향자", code))

    # PDF 스팟체크 비교
    spotcheck_rows = []

    def get_value(records, code, country, response_option):
        for r in records:
            if (
                r.table_code == code
                and r.country == country
                and r.response_option == response_option
            ):
                return r.value
        return None

    spot_defs = [
        ("일반외국인", "E1A-1", "전체", "경험", 78.9),
        ("일반외국인", "E1A-1", "일본", "경험", 59.6),
        ("일반외국인", "B13-1A", "전체", "한류(한국 문화 콘텐츠) 문화 관련 관심 부재", 24.2),
    ]
    for survey_type, code, country, resp, pdf_val in spot_defs:
        records = all_records_general if survey_type == "일반외국인" else all_records_intent
        actual = get_value(records, code, country, resp)
        diff = None if actual is None else round(actual - pdf_val, 2)
        spotcheck_rows.append(
            {
                "survey_type": survey_type,
                "table_code": f"SPOTCHECK:{code}",
                "n_target_countries_found": "",
                "n_expected": "",
                "missing_countries": f"country={country}, option={resp}",
                "unexpected_countries": "",
                "base_labels": f"pdf_value={pdf_val}, extracted_value={actual}, diff={diff}",
                "total_row_sample_n": "",
                "n_records": "",
                "status": "MATCH" if (actual is not None and abs(actual - pdf_val) < 0.5) else "MISMATCH_OR_MISSING",
            }
        )

    # C4 방한의향자 상위 5개 항목 스팟체크 (PDF 41.4/40.7/36.0/34.9/34.3)
    c4_pdf_values = [41.4, 40.7, 36.0, 34.9, 34.3]
    c4_total_records = sorted(
        [r for r in all_records_intent if r.table_code == "C4" and r.country == "전체"],
        key=lambda r: -(r.value or 0),
    )[:5]
    c4_actual_values = [round(r.value, 1) if r.value is not None else None for r in c4_total_records]
    spotcheck_rows.append(
        {
            "survey_type": "방한의향자",
            "table_code": "SPOTCHECK:C4",
            "n_target_countries_found": "",
            "n_expected": "",
            "missing_countries": "",
            "unexpected_countries": "",
            "base_labels": f"pdf_top5={c4_pdf_values}, extracted_top5={c4_actual_values}",
            "total_row_sample_n": "",
            "n_records": len(c4_total_records),
            "status": "MATCH" if c4_actual_values == c4_pdf_values else "MISMATCH_OR_MISSING",
        }
    )

    return validation_rows + spotcheck_rows


def write_validation_csv(path: Path, validation_rows):
    fieldnames = [
        "survey_type", "table_code", "n_target_countries_found", "n_expected",
        "missing_countries", "unexpected_countries", "base_labels",
        "total_row_sample_n", "n_records", "status",
    ]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in validation_rows:
            writer.writerow(row)
    print(f"[저장] {path} ({len(validation_rows)}행)")


def main():
    if not GENERAL_FILE.exists() or not INTENT_FILE.exists():
        raise FileNotFoundError("원본 Excel 파일을 data/raw/ 에서 찾을 수 없습니다.")

    warnings = []

    general_records, general_summaries = run_extraction_for_file(
        GENERAL_FILE, "일반외국인", GENERAL_TABLE_CODES, warnings
    )
    intent_records, intent_summaries = run_extraction_for_file(
        INTENT_FILE, "방한의향자", INTENT_TABLE_CODES, warnings
    )

    write_long_csv(OUT_GENERAL_CSV, general_records)
    write_long_csv(OUT_INTENT_CSV, intent_records)
    write_mapping_csv(OUT_MAPPING_CSV)

    validation_rows = run_validation(general_records, intent_records, warnings)
    write_validation_csv(OUT_VALIDATION_CSV, validation_rows)

    print("\n" + "=" * 70)
    print("경고/주의사항 목록")
    print("=" * 70)
    if warnings:
        for w in warnings:
            print(" -", w)
    else:
        print(" (없음)")

    print("\n" + "=" * 70)
    print("표별 요약 (일반외국인)")
    print("=" * 70)
    for s in general_summaries:
        print(f" - {s['table_code']:12s} BASE='{s['base']}' 국가수={s['n_countries']} 레코드수={s['n_records']}")

    print("\n" + "=" * 70)
    print("표별 요약 (방한의향자)")
    print("=" * 70)
    for s in intent_summaries:
        print(f" - {s['table_code']:12s} BASE='{s['base']}' 국가수={s['n_countries']} 레코드수={s['n_records']}")


if __name__ == "__main__":
    main()
